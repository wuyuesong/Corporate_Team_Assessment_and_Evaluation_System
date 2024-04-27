import hashlib
import uuid
import time
from datetime import datetime

from django.contrib.auth.hashers import make_password, check_password
from django_restql.fields import DynamicSerializerMethodField
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from django.db import connection
from django.db.models import Q
from application import dispatch
from dvadmin.system.models import Users, Role, Dept, Department, EvaluateTask, Task, Staff,EvaluateTaskRank, EvaluateTaskAbnormalData
from dvadmin.system.views.role import RoleSerializer
from dvadmin.utils.json_response import ErrorResponse, DetailResponse, SuccessResponse
from dvadmin.utils.serializers import CustomModelSerializer
from dvadmin.utils.validator import CustomUniqueValidator
from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin.utils.calc import calc_score
from dvadmin.utils.send_email import send_email
import numpy as np

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from urllib.parse import quote



class EvaluateTaskSerializer(CustomModelSerializer):
    """
    用户管理-序列化器
    """

    class Meta:
        model = EvaluateTask
        fields = "__all__"


class EvaluateTaskCreateSerializer(CustomModelSerializer):
    """
    用户新增-序列化器
    """

    def save(self, **kwargs):
        data = super().save(**kwargs)
        data.save()
        return data

    class Meta:
        model = EvaluateTask
        fields = "__all__"


class EvaluateTaskUpdateSerializer(CustomModelSerializer):
    """
    用户修改-序列化器
    """
    def save(self, **kwargs):
        data = super().save(**kwargs)
        data.save()
        return data

    class Meta:
        model = EvaluateTask
        fields = "__all__"
        

class EvaluateTaskInfoUpdateSerializer(CustomModelSerializer):
    """
    用户修改-序列化器
    """

    def update(self, instance, validated_data):
        return super().update(instance, validated_data)

    class Meta:
        model = EvaluateTask
        fields = "__all__"


class ExportEvaluateTaskProfileSerializer(CustomModelSerializer):
    """
    用户导出 序列化器
    """

    class Meta:
        model = EvaluateTask
        fields = "__all__"


class EvaluateTaskProfileImportSerializer(CustomModelSerializer):

    def save(self, **kwargs):
        data = super().save(**kwargs)
        data.save()
        return data

    class Meta:
        model = EvaluateTask
        fields = "__all__"


class EvaluateTaskViewSet(CustomModelViewSet):
    """
    用户接口
    list:查询
    create:新增
    update:修改
    retrieve:单例
    destroy:删除
    """

    queryset = EvaluateTask.objects.all()
    serializer_class = EvaluateTaskSerializer
    create_serializer_class = EvaluateTaskCreateSerializer
    update_serializer_class = EvaluateTaskUpdateSerializer
    # filter_fields = ["name", "username", "gender", "is_active", "dept", "user_type"]
    filter_fields = [
        "evaluate_id",
        "task_weight",
        "evaluated_id",
        "score",
        "grade_complete",
        "grade_date"
    ]
    # search_fields = ["username", "name", "dept__name", "role__name"]
    search_fields = "__all__"
    # 导出
    export_field_label = {
        "evaluate_id":"评价人系统id",
        "task_weight": "任务权重",
        "evaluated_id":"被评价人系统id",
        "score": "分数",
        "grade_complete": "完成情况",
        "grade_date":"评价时间"
    }
    export_serializer_class = ExportEvaluateTaskProfileSerializer
    # 导入
    import_serializer_class = EvaluateTaskProfileImportSerializer

    import_field_dict = {
        "evaluate_id":"评价人系统id",
        "task_weight": "任务权重",
        "evaluated_id":"被评价人系统id",
        "score": "分数",
        "grade_complete": "完成情况",
        "grade_date":"评价时间"
    }

    def evaluate_task_create(self, request: Request):
        task_id = uuid.uuid4()
        task_type = request.data.get("task_type")
        task_name = request.data.get("task_name")
        task_describe = request.data.get("task_describe")
        task_start_date = request.data.get("task_start_date")
        task_start_date = datetime.strptime(task_start_date, '%Y-%m-%d %H:%M:%S')
        task_end_date = request.data.get("task_end_date")
        task_end_date = datetime.strptime(task_end_date, '%Y-%m-%d %H:%M:%S')
        task_create_date = datetime.now()
        Task(task_id=task_id, task_name=task_name, task_describe=task_describe, task_start_date=task_start_date, task_end_date=task_end_date, task_create_date=task_create_date, task_type=0).save()
        evaluate = request.data.get("evaluate")
        evaluated = request.data.get("evaluated")
        # time1 = time.time()
        tmp_list =[]
        for evaluate_one in evaluate:
            for evaluated_one in evaluated:
                tmp_list.append(EvaluateTask(task_id=task_id, evaluate_id=evaluate_one["evaluate_id"], task_weight=evaluate_one["task_weight"],evaluated_id=evaluated_one["evaluated_id"]))
        EvaluateTask.objects.bulk_create(tmp_list)
        # time2 = time.time()
        # print("生成任务耗时：", time2 - time1)

        return DetailResponse(data=dict(task_id=task_id), msg="创建成功")

    @action(methods=['POST'], detail=False, permission_classes=[])
    def evaluate_task_info(self, request: Request):
        task_id = request.data.get("task_id")
        staff_id = request.data.get("staff_id")
        evaluated_id_list = list(EvaluateTask.objects.filter(task_id=task_id, evaluate_id=staff_id).values_list('evaluated_id', flat=True).distinct().order_by('evaluated_id'))
        evaluated_list = EvaluateTask.objects.filter(task_id=task_id, evaluate_id=staff_id)
        evaluated_queryset = Staff.objects.filter(staff_id__in=evaluated_id_list)
        ret = []
        score_dict = {}
        for evaluated in evaluated_list:
            score_dict[evaluated.evaluated_id] = evaluated.score
        for evaluated in evaluated_queryset:
            ret.append(dict(evaluated_id=evaluated.staff_id, evaluated_name=evaluated.staff_name, evaluated_rank=evaluated.staff_rank, evaluated_department=evaluated.staff_department, score=score_dict[evaluated.staff_id]))

        return DetailResponse(data=ret, msg="查询成功")
    
    @action(methods=['POST'], detail=False, permission_classes=[])
    def submit_evaluate_task(self, request: Request):
        evaluate_id = request.data.get("evaluate_id")
        task_id = request.data.get("task_id")
        scores = request.data.get("scores")
        submit_type = request.data.get("submit_type")
        
        for score in scores:
            evaluated_id = score["evaluated_id"]
            if submit_type == 1:
                EvaluateTask.objects.filter(evaluate_id=evaluate_id, task_id=task_id, evaluated_id=evaluated_id).update(score=score["score"], grade_complete=1, grade_date=datetime.now())
            else:
                EvaluateTask.objects.filter(evaluate_id=evaluate_id, task_id=task_id, evaluated_id=evaluated_id).update(score=score["score"], grade_date=datetime.now())

        return DetailResponse(data=[], msg="提交成功")


    def evaluate_task_delete_all(self, request: Request):
        EvaluateTask_all = EvaluateTask.objects.all()
        EvaluateTask_all.delete()
        return DetailResponse(data=[], msg="删除成功")
    

    @action(methods=['POST'], detail=False, permission_classes=[])
    def task_info(self, request: Request):
        task_id = request.data.get("task_id")
        task = Task.objects.get(task_id=task_id)

        all_evaluate = list(EvaluateTask.objects.filter(task_id=task_id).values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        undo_staff_id_list = list(EvaluateTask.objects.filter(task_id=task_id, grade_complete=0).values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        undo_staff_info = Staff.objects.filter(staff_id__in=undo_staff_id_list)

        ret = {"task_id":task.task_id,
                "task_name":task.task_name,
                "task_describe":task.task_describe,
                "task_start_date":task.task_start_date,
                "task_end_date":task.task_end_date,
                "task_create_date":task.task_create_date,
                "task_done":task.task_done,
                "staff_count":len(all_evaluate),}
        undo_staff = []
        for staff in undo_staff_info:
            undo_staff.append({
                "staff_name":staff.staff_name,
                "staff_firm_id":staff.staff_id,
                "staff_department":staff.staff_department
            })

        ret["undo_staff"] = undo_staff

        return DetailResponse(data=ret, msg="获取成功")     
    

    @action(methods=['post'], detail=False, permission_classes=[])
    def task_delete_single(self, request: Request):
        task_id = request.data.get("task_id")
        task_all = Task.objects.filter(task_id=task_id)
        evaluateTask_all = EvaluateTask.objects.filter(task_id=task_id)
        task_all.delete()
        evaluateTask_all.delete()
        return DetailResponse(data=[], msg="删除成功")
    
    @action(methods=['post'], detail=False, permission_classes=[])
    def task_calc(self, request: Request):
        task_id = request.data.get("task_id")
        mul=request.data.get("mul")
        
        task_all = EvaluateTask.objects.filter(task_id=task_id)

        all_evaluate = list(EvaluateTask.objects.filter(task_id=task_id).values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        all_evaluated = list(EvaluateTask.objects.filter(task_id=task_id).values_list('evaluated_id', flat=True).distinct().order_by('evaluated_id'))

        map_evaluate = {}
        map_evaluated = {}
        for index, evaluate_id in enumerate(all_evaluate):
            map_evaluate[evaluate_id] = index
        
        for index, evaluated_id in enumerate(all_evaluated):
            map_evaluated[evaluated_id] = index

        scores = np.zeros((len(all_evaluate), len(all_evaluated)))
        weight = np.array(len(all_evaluate) * [0])
        for task in task_all:
            i = map_evaluate[task.evaluate_id]
            j = map_evaluated[task.evaluated_id]
            scores[i, j] = task.score
            weight[i] = task.task_weight
        ranks, abnormal_datas = calc_score(len(all_evaluate), len(all_evaluated), mul, np.array(all_evaluated), np.array(all_evaluate), scores, weight)
        for rank in ranks:
            EvaluateTaskRank.objects.create(task_id=task_id, evaluated_id=rank["id"], evaluated_rank=rank["rank"], evaluated_score=rank["score"])

        for abnormal_data in abnormal_datas:
            EvaluateTaskAbnormalData.objects.create(task_id=task_id, evaluate_id=abnormal_data["evaluate_id"], evaluated_id=abnormal_data["evaluated_id"],origin_value=abnormal_data["origin_value"],fix_value=abnormal_data["fix_value"])

        Task.objects.filter(task_id=task_id).update(task_done=1)

        return DetailResponse(data=[], msg="计算成功")
    
    @action(methods=['post'], detail=False, permission_classes=[])
    def get_rank(self, request: Request):
        task_id = request.data.get("task_id")

        rank_list = EvaluateTaskRank.objects.filter(task_id=task_id)

        ret= []
        for rank in rank_list:
            evaluated_name = Staff.objects.get(staff_id=rank.evaluated_id).staff_name
            ret.append(dict(evaluated_id=rank.evaluated_id, evaluated_rank=rank.evaluated_rank, evaluated_score=rank.evaluated_score, evaluated_name=evaluated_name))

        return DetailResponse(data=ret, msg="获取成功")
    
    @action(methods=['post'], detail=False, permission_classes=[])
    def get_abnormal_data(self, request: Request):
        task_id = request.data.get("task_id")

        abnormal_data_list = EvaluateTaskAbnormalData.objects.filter(task_id=task_id)

        ret= []
        for abnormal_data in abnormal_data_list:
            evaluated_name = Staff.objects.get(staff_id=abnormal_data.evaluated_id).staff_name
            evaluate_name = Staff.objects.get(staff_id=abnormal_data.evaluate_id).staff_name
            ret.append(dict(evaluate_id=abnormal_data.evaluate_id, 
                            evaluated_id=abnormal_data.evaluated_id, 
                            origin_value=abnormal_data.origin_value, 
                            fix_value=abnormal_data.fix_value, 
                            evaluated_name=evaluated_name,
                            evaluate_name=evaluate_name,))

        return DetailResponse(data=ret, msg="获取成功")
    
    @action(methods=['post'], detail=False, permission_classes=[])
    def reset_taskres(self, request:Request):
        task_id = request.data.get("task_id")
        if not task_id:
            return DetailResponse(msg="task_id不能为空")
        # 清空rank和abnoraml记录
        EvaluateTaskRank.objects.filter(task_id=task_id).delete()
        EvaluateTaskAbnormalData.objects.filter(task_id=task_id).delete()
        # 并将任务状态设置回0
        Task.objects.filter(task_id=task_id).update(task_done=0)
        return DetailResponse(msg="重置成功",data=[])
    

    @action(methods=['post'], detail=False, permission_classes=[])
    def get_all_evaluate(self, request: Request):
        task_id = request.data.get("task_id")
        all_evaluate_id = list(EvaluateTask.objects.filter(task_id=task_id).values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        staff_infos = Staff.objects.filter(staff_id__in=all_evaluate_id)

        ret = []
        for staff_info in staff_infos:
            ret.append({
                "staff_name":staff_info.staff_name,
                "staff_firm_id":staff_info.staff_id,
                "staff_department":staff_info.staff_department,
                "staff_id":staff_info.staff_id
            })

        return DetailResponse(data=ret, msg="获取成功")
    

    @action(methods=['get'], detail=False, permission_classes=[])
    def send_email(self, request: Request):
        all_evaluate_id = list(EvaluateTask.objects.all().values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        staff_infos = Staff.objects.filter(staff_id__in=all_evaluate_id)

        to_addrs = []
        ret = []
        for staff_info in staff_infos:
            to_addrs.append({
                "staff_name":staff_info.staff_name,
                "addr":staff_info.staff_email,
                "username": staff_info.username,
                "password": staff_info.password
            })

        send_email(to_addrs=to_addrs)
        
        return DetailResponse(data=[], msg="发送成功")
    
    @action(methods=['get'], detail=False, permission_classes=[])
    def generate_excel(self, request: Request):
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出员工账号密码.xlsx"))}'
        wb = Workbook()
        ws1 = wb.create_sheet("data", 1)
        ws1.sheet_state = "hidden"
        ws = wb.active

        all_evaluate_id = list(EvaluateTask.objects.all().values_list('evaluate_id', flat=True).distinct().order_by('evaluate_id'))
        staff_infos = Staff.objects.filter(staff_id__in=all_evaluate_id)

        data = np.zeros((len(all_evaluate_id), 5))

        for index, staff in enumerate(staff_infos):
            data[index][0] = staff.staff_id
            data[index][1] = staff.password
            data[index][2] = staff.staff_department
            data[index][3] = staff.staff_rank
            data[index][4] = list(EvaluateTask.objects.filter(evaluate_id=staff.staff_id).values_list('task_id', flat=True).distinct().order_by('task_id'))

        idex=np.lexsort([data[:,4], data[:,3], data[:,2]])

        sorted_data = data[idex, :]
         
        for index, data in enumerate(sorted_data):
            if index == 0:
                ws.append(data)
            else:
                if sorted_data[index][2] != sorted_data[index-1][2] or sorted_data[index][3] != sorted_data[index-1][3] or sorted_data[index][4] != sorted_data[index-1][4]:
                    ws.append([])
                ws.append(data)
            
        tab = Table(displayName="Table")  # 名称管理器
        ws.add_table(tab)
        wb.save(response)
        return response
        


    
        




